import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font
import logging

# 配置日志
logging.basicConfig(filename='process_clauses.log', level=logging.ERROR,
                    format='%(asctime)s - %(levelname)s - %(message)s')

def parse_clause(clause_str):
    """解析条款字符串，返回条、款、项、目，处理空格和格式不一致"""
    if not clause_str or pd.isna(clause_str):
        return None, None, None, None
    
    clause_str = str(clause_str).replace(".1", "").replace(" ", "").strip()
    if "《" in clause_str and "》" in clause_str:
        clause_str = clause_str.split("》")[1]
    
    parts = clause_str.split("第")
    strip, kuan, xiang, mu = None, None, None, None
    
    for part in parts[1:]:
        if part.endswith("条"):
            strip = part[:-1]
        elif part.endswith("款"):
            kuan = part[:-1]
        elif part.endswith("项"):
            xiang = part[:-1]
        elif part.endswith("目"):
            mu = part[:-1]
    
    strip = int(strip) if strip else 0
    kuan = int(kuan) if kuan else 0
    xiang = int(xiang) if xiang else 0
    mu = int(mu) if mu else 0
    
    return strip, kuan, xiang, mu

def find_content_in_law_structure(strip, kuan, xiang, mu, law_df):
    if xiang != 0 or mu != 0:
        query = "条 == @strip and 款 == @kuan and 项 == @xiang and 目 == @mu"
    elif kuan != 0:
        query = "条 == @strip and 款 == @kuan and 项 == 0 and 目 == 0"
    else:
        return None
    
    results = law_df.query(query)
    if results.empty:
        return None
    if strip and strip > 10000:
        return results["内容"].iloc[0]
    else:
        return results["原内容"].iloc[0]
    #penalty_results = results[results["可能违则"] == 1]
     #说明查找的是违则
    #if not penalty_results.empty:
    #    return penalty_results.sort_values(by="可能度", ascending=False)["内容"].iloc[0]
    
    #return results.sort_values(by="可能度", ascending=False)["原内容"].iloc[0]

def find_related_law(strip, kuan, xiang, mu, law_df):
    """查找条款对应的关联法律"""
    query = "条 == @strip and 款 == @kuan and 项 == @xiang and 目 == @mu"
    results = law_df.query(query)
    if results.empty:
        return None
    return results.iloc[0]["关联法律"] if "关联法律" in results.columns and not pd.isna(results.iloc[0]["关联法律"]) else None

def format_clause(strip, kuan, xiang, mu, related_law=None):
    """格式化条款字符串，支持带关联法律"""
    clause = f"第{strip}条"
    if kuan:
        clause += f"第{kuan}款"
    if xiang:
        clause += f"第{xiang}项"
    if mu:
        clause += f"第{mu}目"
    if related_law:
        clause = f"《{related_law}》{clause}"
    return clause

def process_clauses(clauses_str, law_structure_df):
    """通用函数：处理条款字符串（罚则或违则），返回内容和更新后的条款列表"""
    try:
        clauses = str(clauses_str).split("|")
        contents = []
        updated_clauses = []
        
        for clause in clauses:
            strip, kuan, xiang, mu = parse_clause(clause)
            if strip and strip > 10000:
                related_law = find_related_law(strip, kuan, xiang, mu, law_structure_df)
                adjusted_strip = strip - (strip // 10000) * 10000
                updated_clause = format_clause(adjusted_strip, kuan, xiang, mu, related_law)
            else:
                updated_clause = clause
            content = find_content_in_law_structure(strip, kuan, xiang, mu,law_structure_df)
            if content:
                contents.append(content)
            updated_clauses.append(updated_clause)
        
        return "|".join(contents), "|".join(updated_clauses)
    except Exception as e:
        # 记录错误信息到日志文件
        print(f"Error processing clauses: {clauses_str}. Error message: {e}")
        return None, None

def update_law_result(input_file, law_structure_file, output_file):
    """更新 LawAIResult 文件，处理条款内容并调整格式"""
    wb_law_structure = openpyxl.load_workbook(law_structure_file)
    law_structure_sheets = {sheet.title: pd.read_excel(law_structure_file, sheet_name=sheet.title) 
                           for sheet in wb_law_structure.worksheets}

    wb_input = openpyxl.load_workbook(input_file)
    wb_output = openpyxl.Workbook()
    
    for sheet_name in wb_input.sheetnames:
        if sheet_name not in law_structure_sheets:
            print(f"Skipping sheet {sheet_name}: no matching sheet in {law_structure_file}")
            continue

        law_ai_df = pd.read_excel(input_file, sheet_name=sheet_name)
        required_columns = ["事由", "行为", "罚则", "违则", "说明"]
        if not all(col in law_ai_df.columns for col in required_columns):
            print(f"Skipping sheet {sheet_name}: missing required columns")
            continue

        law_structure_df = law_structure_sheets[sheet_name]
        law_ai_df["违法行为"] = None
        law_ai_df["罚则条款"] = None
        law_ai_df["违则条款"] = None

        for index, row in law_ai_df.iterrows():
            # 处理违法行为
            strip, kuan, xiang, mu = parse_clause(str(row["行为"]))
            content = find_content_in_law_structure(strip, kuan, xiang, mu,law_structure_df)
            law_ai_df.at[index, "违法行为"] = content

            # 处理罚则条款
            penalty_contents, updated_penalty_clauses = process_clauses(row["罚则"], law_structure_df)
            law_ai_df.at[index, "罚则条款"] = penalty_contents
            law_ai_df.at[index, "罚则"] = updated_penalty_clauses

            # 处理违则条款
            violation_contents, updated_violation_clauses = process_clauses(row["违则"], law_structure_df)
            law_ai_df.at[index, "违则条款"] = violation_contents
            law_ai_df.at[index, "违则"] = updated_violation_clauses

        # 调整列顺序
        new_order = ["编号", "事由", "说明", "违则", "违则条款", "行为", "违法行为", "罚则", "罚则条款"]
        law_ai_df = law_ai_df[new_order]

        # 写入新 Sheet
        if sheet_name in wb_output.sheetnames:
            ws = wb_output[sheet_name]
        else:
            ws = wb_output.create_sheet(title=sheet_name)
        
        for c_idx, col in enumerate(law_ai_df.columns, 1):
            ws.cell(row=1, column=c_idx, value=col)
        
        for r_idx, row in law_ai_df.iterrows():
            for c_idx, col in enumerate(law_ai_df.columns, 1):
                ws.cell(row=r_idx + 2, column=c_idx, value=row[col])

        # 设置列宽和格式
        column_widths = {
            "行为": 7, "罚则": 7, "违则": 7, "说明": 20, "事由": 20, "违法行为": 20,
        }
        for col in ws.columns:
            header = col[0].value
            width = column_widths.get(header, 40) * 1.2
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = width

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        bold_font = Font(bold=True, size=12)

        for row in ws.rows:
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                cell.border = thin_border
                header = ws.cell(row=1, column=cell.column).value
                if header in ["行为", "罚则", "违则"]:
                    cell.font = bold_font

    if "Sheet" in wb_output.sheetnames:
        wb_output.remove(wb_output["Sheet"])
    wb_output.save(output_file)

if __name__ == "__main__":
    input_file = "result\\LawAIResult.xlsx"
    law_structure_file = "result\\law_structure_format_num_ex.xlsx"
    output_file = "result\\law_cause.xlsx"
    update_law_result(input_file, law_structure_file, output_file)
    print(f"Updated file saved as {output_file}")