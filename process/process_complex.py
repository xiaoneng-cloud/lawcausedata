import pandas as pd
import re
import os
from openpyxl import load_workbook

def process_excel(file_path):
    # 读取 Excel 文件
    xls = pd.ExcelFile(file_path)
    # 加载工作簿用于写入数据
    wb = load_workbook(file_path)

    # 遍历所有 sheet
    for sheet_name in xls.sheet_names:
        # 读取当前 sheet
        df = xls.parse(sheet_name)

        # 确保必要的列存在
        required_columns = ['条', '款', '项', '目', '内容', '罚则', '违法情形']
        if not all(col in df.columns for col in required_columns):
            print(f"Sheet {sheet_name} 缺少必要的列，跳过处理")
            continue

        # 处理每一行数据
        for idx, row in df.iterrows():
            content = row['内容']
            current_tiao = row['条']
            current_kuan = row['款'] if not pd.isna(row['款']) else 0
            current_xiang = row['项'] if not pd.isna(row['项']) else 0
            current_mu = row['目'] if not pd.isna(row['目']) else 0

            # 步骤 6：处理罚则和违法情形
            if row['罚则'] == 1 and row['违法情形'] == 0:
                matches = re.findall(r'第(\d+)条(?!第\d+款)', content)
                for match in matches:
                    tiao_num = int(match)
                    relevant_rows = df[df['条'] == tiao_num]
                    penalty_clauses = []
                    for _, rel_row in relevant_rows.iterrows():
                        if rel_row['罚则'] == 1:
                            clause = f"第{tiao_num}条"
                            if rel_row['款'] > 0:
                                clause += f"第{int(rel_row['款'])}款"
                            if rel_row['项'] > 0:
                                clause += f"第{int(rel_row['项'])}项"
                            penalty_clauses.append(clause)
                    if penalty_clauses:
                        replacement = '、'.join(penalty_clauses)
                        content = content.replace(f"第{tiao_num}条", replacement)

            elif row['罚则'] == 0 and row['违法情形'] == 1:
                matches = re.findall(r'第(\d+)条(?!第\d+款)', content)
                for match in matches:
                    tiao_num = int(match)
                    relevant_rows = df[df['条'] == tiao_num]
                    violation_clauses = []
                    for _, rel_row in relevant_rows.iterrows():
                        if rel_row['违法情形'] == 1:
                            clause = f"第{tiao_num}条"
                            if rel_row['款'] > 0:
                                clause += f"第{int(rel_row['款'])}款"
                            if rel_row['项'] > 0:
                                clause += f"第{int(rel_row['项'])}项"
                            violation_clauses.append(clause)
                    if violation_clauses:
                        replacement = '、'.join(violation_clauses)
                        content = content.replace(f"第{tiao_num}条", replacement)

            # 更新处理后的内容
            df.at[idx, '内容'] = content

        # 获取当前 sheet 的对象
        ws = wb[sheet_name]
        # 写入表头
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
        # 写入数据
        for row_idx, row in df.iterrows():
            for col_idx, col_name in enumerate(df.columns, start=1):
                ws.cell(row=row_idx + 2, column=col_idx, value=row[col_name])

    # 保存修改后的工作簿
    wb.save(file_path)
    print(f"处理完成，结果已保存至: {file_path}")

if __name__ == "__main__":
    file_path = 'result\\law_structure_format_num_ex.xlsx'
    if os.path.exists(file_path):
        process_excel(file_path)
    else:
        print(f"错误：文件 {file_path} 不存在。")