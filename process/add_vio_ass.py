import pandas as pd
from openpyxl import load_workbook
import re
import jieba
from gensim.models import KeyedVectors
from tools.extract_violation_articles import PENALTY_TYPES
import numpy as np
from tools.extract_violation_articles import get_violation_pattern
from tools.extract_violation_articles import get_clause_array
from tools.extract_violation_articles import find_associated_body


# 定义辅助词和否定词，增加 "下列行为" 和 "之一"，用于语义比较
REMOVE_WORDS = ["不", "未", "的", "下列行为", "之一"]


def load_pretrained_model():
    try:
        model = KeyedVectors.load_word2vec_format('AILab.bin', binary=True)
        return model
    except FileNotFoundError:
        print("未找到预训练模型文件，请检查文件路径。")
        return None


def preprocess(sentence):
    """
    预处理句子，去除特殊词并分词
    """
    for word in REMOVE_WORDS:
        sentence = sentence.replace(word, "")
    return jieba.lcut(sentence)


def get_sentence_vector(sentence, model):
    """
    根据预训练的 Word2Vec 模型获取句子向量
    """
    vectors = []
    for word in sentence:
        if word in model:
            vectors.append(model[word])
    if not vectors:
        return np.zeros(model.vector_size)
    return np.mean(vectors, axis=0)


def cosine_similarity(vec1, vec2):
    """
    计算两个向量的余弦相似度
    """
    dot_product = np.dot(vec1, vec2)
    norm_vec1 = np.linalg.norm(vec1)
    norm_vec2 = np.linalg.norm(vec2)
    if norm_vec1 == 0 or norm_vec2 == 0:
        return 0
    return dot_product / (norm_vec1 * norm_vec2)


# 判断是否包含（基于 Word2Vec 余弦相似度）
def is_mostly_contained(sentence1, sentence2, model, threshold=0.85):
    # 预处理两句话
    words1 = preprocess(sentence1)
    words2 = preprocess(sentence2)

    # 获取句子向量
    vec1 = get_sentence_vector(words1, model)
    vec2 = get_sentence_vector(words2, model)

    # 计算余弦相似度
    similarity = cosine_similarity(vec1, vec2)

    return similarity >= threshold, similarity


def extract_violation_clauses(text, df):  # 添加 df 参数，用于获取数据
    """
    从文本中抽取符合格式的内容
    格式要求：(?:未依照|违反|有)必须有一个，当前面为"有"时，后面必须有"致使"
    返回格式：第X条|第Y条第Z款
    """
    # 定义正则表达式模式
    pattern = get_violation_pattern()
    match = re.search(pattern, text)
    #print(f"匹配text：{text}")
    
    if  match:
        result_parts = []
        #print(f"找到匹配：{match.group()}")
        # Extract clauses using get_clause_array
        clause_array = get_clause_array(match.group())
        #print(f"解析后的条款数组：{clause_array}")
        
        # Convert each clause object to text format
        for clause in clause_array:
            # 使用f-string避免字符串拼接
            clause_text = f"第{clause['Article']}条"
            if clause.get('Section'):
                clause_text += f"第{clause['Section']}款"
            if clause.get('Item'):
                clause_text += f"第{clause['Item']}项"
            if clause.get('Subitem'):
                clause_text += f"第{clause['Subitem']}目"
            if re.match(r'第\d+条$', clause_text):  # 判断是否为“第X条”格式
                article_num = int(re.findall(r'\d+', clause_text)[0])
                # 查找同条下“违法情形”为1的记录
                same_article_records = df[(df['条'] == article_num) & (df['违法情形'] == 1)]
                if not same_article_records.empty:
                    for _, row in same_article_records.iterrows():
                        new_clause = format_clause(row)
                        result_parts.append(new_clause)
                else:
                    result_parts.append(clause_text)
            else:
                result_parts.append(clause_text)
            
                #print(f"生成的条款文本：{clause_text}")
            
        result = "|".join(result_parts)
    else:
        return ""
    # Join all parts with "|"
    #print(f"最终结果：{result}")
    return result


def format_clause(row):
    """
    格式化条、款、项、目为合适的结构
    """
    clause_str = f"第{int(row['条'])}条"
    if row['款']:
        clause_str += f"第{int(row['款'])}款"
    if row['项']:
        clause_str += f"第{int(row['项'])}项"
    if row['目']:
        clause_str += f"第{int(row['目'])}目"
    return clause_str


def get_superior_text(df, current_row):
    if bool(current_row['项']) and not bool(current_row['目']):
        superior_conditions = (df['条'] == current_row['条']) & (df['款'] == current_row['款']) & (df['项'] == 0)
        superior_rows = df[superior_conditions]
        if not superior_rows.empty:
            superior_row = superior_rows.iloc[0]
            parts = re.split(r'[;,，；]', superior_row['内容'])
            for part in parts:
                if "下列" in part:
                    return part + current_row['内容']
    return current_row['内容']




def update_excel(file_path, model):
    try:
        wb = load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            print(f"Processing sheet: {sheet_name}")
            ws = wb[sheet_name]
            
            # Read data from current sheet
            data = []
            headers = None
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    headers = list(row)  # Save headers
                else:
                    data.append(row)
                    
            # Skip empty sheets
            if not headers or not data:
                print(f"Sheet {sheet_name} is empty, skipping...")
                continue
                
            # Create DataFrame with the correct headers
            df = pd.DataFrame(data, columns=headers)

            # 将条、款、项、目列的数据转换为整数类型，空值转换为 0
            for col in ['条', '款', '项', '目']:
                if col in df.columns:
                    df[col] = df[col].fillna(0).astype(int)
                else:
                    print(f"Warning: Column '{col}' not found in sheet {sheet_name}")

            # 删除已有的可能违则和可能度列
            if '可能违则' in df.columns:
                df = df.drop(columns='可能违则')
            if '可能度' in df.columns:
                df = df.drop(columns='可能度')

            # 添加新的列
            df['可能违则'] = ""
            df['可能度'] = ""

            # 处理罚则为 0 且违法情形为 1 的记录
            valid_rows = df[(df['罚则'] == 0) & (df['违法情形'] == 1)].index
            #print(f"Found {len(valid_rows)} rows with 罚则=0 and 违法情形=1 in sheet {sheet_name}")
            
            for index in valid_rows:
                row = df.loc[index]
                violation_clauses = extract_violation_clauses(row['内容'], df)  # 传递 df 参数
                if violation_clauses:
                    df.at[index, '可能违则'] = violation_clauses
                    df.at[index, '可能度'] = 1
                    df.at[index, '条类型'] = 31
                    continue

                if row['项'] > 0:
                    same_kuan_first_row = df[(df['条'] == row['条']) & (df['款'] == row['款']) & (df['项'] == 0)].iloc[0]
                    if same_kuan_first_row['条类型'] == 6:
                        B_text = same_kuan_first_row['内容']
                        new_violation_clauses = extract_violation_clauses(B_text, df)
                        if new_violation_clauses:
                            df.at[index, '可能违则'] = new_violation_clauses
                            df.at[index, '可能度'] = 1
                            df.at[index, '条类型'] = 311
                            continue

                current_text = get_superior_text(df, row)
                current_article = row['条']
                other_records = df[(df['条'] != current_article) & (df['条'] < 10000) & ~((df['罚则'] == 1) | (df['违法情形'] == 1))]

                # 遍历其他记录，检查是否与当前记录相似
                matching_clauses = []
                matching_similarities = []
                for other_index, other_row in other_records.iterrows():
                    is_contained, similarity = is_mostly_contained(current_text, other_row['内容'], model)
                    if is_contained and similarity > 0.9:
                        matching_clauses.append(format_clause(other_row))
                        matching_similarities.append(similarity)
                
                # 对匹配结果进行排序
                if matching_clauses:
                    sorted_indices = np.argsort(matching_similarities)[::-1][:5]
                    top_matching_clauses = [matching_clauses[i] for i in sorted_indices]
                    top_matching_similarities = [matching_similarities[i] for i in sorted_indices]

                    df.at[index, '可能违则'] = "|".join(top_matching_clauses)
                    df.at[index, '可能度'] = "|".join(map(str, top_matching_similarities))
                    df.at[index, '条类型'] = 32

            # 处理罚则为 1 且违法情形为 1 的记录，对应类型4
            valid_rows = df[(df['罚则'] == 1) & (df['违法情形'] == 1)].index
            #print(f"Found {len(valid_rows)} rows with 罚则=1 and 违法情形=1 in sheet {sheet_name}")
            
            for index in valid_rows:
                row = df.loc[index]
                # 提取关联的违则的主体部分，规则是在罚则词之前或者在情形描述之前
                associated_body = find_associated_body(row['内容'])            
                # 需要进一步判断associated_body是否是一个合格的违法行为的描述
                pattern = get_violation_pattern()
                match = re.search(pattern, associated_body)
                 # 增加判断第一个逗号前是否为“的”的条件
                comma_index = associated_body.find('，')
                first_part_ends_with_de = comma_index > 0 and associated_body[comma_index - 1] == '的'

                # 增加判断：associated_body最后一个的文字是不是“，”，同时判断其前面是不是“的”
                last_comma_index = len(associated_body) - 1
                last_part_ends_with_de = last_comma_index > 0 and associated_body[last_comma_index] == '，' and associated_body[last_comma_index - 1] == '的'
                #print(f"associated_body1 {associated_body}")
                if match or first_part_ends_with_de or last_part_ends_with_de:
                    # 抽取有直接描述的条款号
                    violation_clauses = extract_violation_clauses(associated_body.replace(" ", ""), df)  # 传递 df 参数
                    
                    if violation_clauses:
                         # 拆分可能违则的条款号
                        #针对以下类型特殊处理，其前款为父子结构：有前款第（一）项、第（五）项行为之一的，
                        #处2万元以上10万元以下的罚款；有前款第（二）项、第（三）项行为的，处2万元以下的罚款；有前款第（四）项行为的，处5万元以上20万元以下的罚款。
                        clause_list = violation_clauses.split('|')
                        is_type_3_or_311 = False
                        additional_penalty = ""
                        for clause in clause_list:
                            # 解析条款号获取条、款、项、目
                            match = re.match(r'第(\d+)条(?:第(\d+)款)?(?:第(\d+)项)?(?:第(\d+)目)?', clause)
                            if match:
                                article = int(match.group(1))
                                section = int(match.group(2)) if match.group(2) else None
                                item = int(match.group(3)) if match.group(3) else None
                                subitem = int(match.group(4)) if match.group(4) else None
                                
                                # 筛选出符合条件的记录
                                condition = (df['条'] == article)
                                if section is not None:
                                    condition &= (df['款'] == section)
                                if item is not None:
                                    condition &= (df['项'] == item)
                                if subitem is not None:
                                    condition &= (df['目'] == subitem)
                                
                                clause_rows = df[condition]
                                
                                for _, clause_row in clause_rows.iterrows():
                                    print(f"clause_rows {clause_row['条款'] }")
                                    if clause_row['条类型'] in [3, 311]:
                                        is_type_3_or_311 = True
                                        # 查找同条同款类型为6的记录
                                        condition = (df['条'] == article) & (df['款'] == section) & (df['条类型'] == 6)
                                        if not condition.empty:
                                            df.loc[condition, '增加处罚'] = df.at[index, '条款']
                        if is_type_3_or_311:
                            df.at[index, '条类型'] = 5
                            df.at[index, '违法情形'] = 0
                            continue
                        else:
                            # 继续原逻辑
                            df.at[index, '可能违则'] = violation_clauses
                            df.at[index, '可能度'] = 1
                            df.at[index, '条类型'] = 41
                            continue
                        
                    # 遍历其他记录，检查是否与当前记录相似
                    #associated_body 要进行进一步的处理
                    # 新增逻辑：从后往前找第一个“的，”，获取其前面的文字，这样对比更为准确
                    #print(f"associated_body1 {associated_body}")
                    pos = associated_body.rfind('的，')
                    if pos != -1:
                        associated_body = associated_body[:pos]
                        df.at[index, '违法行为文本'] = associated_body
                    #print(f"associated_body2 {associated_body}")
                    current_article = row['条']
                    other_records = df[(df['条'] != current_article) & (df['条'] < 10000) & ~((df['罚则'] == 1) | (df['违法情形'] == 1))]
                    matching_clauses = []
                    matching_similarities = []
                    for _, other_row in other_records.iterrows():
                        is_contained, similarity = is_mostly_contained(associated_body, other_row['内容'], model)
                        if is_contained and similarity > 0.9:
                            matching_clauses.append(format_clause(other_row))
                            matching_similarities.append(similarity)

                    if matching_clauses:
                        sorted_indices = np.argsort(matching_similarities)[::-1][:5]
                        top_matching_clauses = [matching_clauses[i] for i in sorted_indices]
                        top_matching_similarities = [matching_similarities[i] for i in sorted_indices]

                        df.at[index, '可能违则'] = "|".join(top_matching_clauses)
                        df.at[index, '可能度'] = "|".join(map(str, top_matching_similarities))
                        df.at[index, '条类型'] = 44

            # 获取最新的列名（考虑可能添加了新列）
            updated_headers = df.columns.tolist()
            
            # 清空原工作表内容（保留表头行）
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.value = None
            
            # 如果需要，先删除多余的行
            if ws.max_row > len(df) + 1:  # +1 是因为有表头行
                ws.delete_rows(len(df) + 2, ws.max_row - len(df) - 1)
            
            # 更新表头以包含新列
            for col_idx, header in enumerate(updated_headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            
            # 逐行写入数据
            for row_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    column_name = updated_headers[col_idx - 1]
                    # 处理不同类型的值
                    if isinstance(value, (list, tuple)):
                        value = "|".join(map(str, value))
                    elif pd.isna(value):  # 处理 NaN 值
                        value = ""
                    elif isinstance(value, (int, float)) and column_name not in ['可能度']:
                        # 对于非可能度列的数值，保持为数值类型
                        pass
                    else:
                        value = str(value)
                    
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            #print(f"Finished processing sheet: {sheet_name}, wrote {len(df)} rows")

        # 保存工作簿
        wb.save(file_path)
        print("File updated successfully.")
    except Exception as e:
        import traceback
        print(f"Error updating Excel file: {e}")
        print(traceback.format_exc())  # 打印详细的错误堆栈

if __name__ == "__main__":
    model = load_pretrained_model()
    if model is None:
        exit(1)
    file_path = "result\\law_structure_format_num_ex.xlsx"
    update_excel(file_path, model)