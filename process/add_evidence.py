import pandas as pd
from openpyxl import load_workbook
import re
from process.tools.extract_violation_articles import get_violation_pattern
from process.tools.extract_violation_articles import find_associated_body


# 定义行政处罚相关的正则表达式模式
PUNISHMENT_PATTERNS = [
    re.compile(r'(?!.*行政处分)罚款'),
    re.compile(r'(?!.*行政处分)没收(违法所得|非法财物)'),
    re.compile(r'(?!.*行政处分)责令(改正|停产停业|停止|关闭)'),
    re.compile(r'(?!.*行政处分)暂扣|吊销(?!\s*(处|罚))(.*(许可证|执照|资格证书|资质证书|驾驶证|驾照|营业执照))'),
    re.compile(r'(?!.*行政处分)拘留|强制报废|强制排除妨碍'),
    re.compile(r'(?!.*行政处分)降低资质等级'),
    re.compile(r'(?!.*行政处分)限制从业'),
    # 新增的正则表达式模式
    re.compile(r'(?!.*行政处分)(违反|未依照).*?(条(?!例)|款|项)(?!.*治安).*?处罚')
]


def check_discretion(content):
    """
    检查内容中是否包含行政处罚关键词
    :param content: 法律条文内容
    :return: 包含返回 1，不包含返回 0
    """
    for pattern in PUNISHMENT_PATTERNS:
        match = pattern.search(content)
        if match:
            return 1
    return 0


def find_children(df, row):
    """
    查找当前行的子节点
    :param df: 数据框
    :param row: 当前行
    :return: 子节点数据框
    """
    article = row['条']
    paragraph = row['款']
    item = row['项']
    # 额外检查和转换 item
    if pd.isna(item):
        item = None
    #print(f"item {item} ")
    if paragraph:
        if item:
            # 如果有款和项，查找目
            condition = (df['条'] == article) & (df['款'] == paragraph) & (df['项'] == item) & (df['目'].notnull())
            #print(f"condition1 {condition} ")
        else:
            # 如果只有款，查找项
            condition = (df['条'] == article) & (df['款'] == paragraph) & (df['项'].notnull())
            #print(f"condition2 {condition} ")
    else:
        # 如果没有款，查找款
        condition = (df['条'] == article) & (df['款'].notnull())
        #print(f"condition3 {condition} ")

    return df[condition]


def update_excel(file_path):
    wb = load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        df = pd.DataFrame(data[1:], columns=data[0])

        # 先转换数据类型
        df = df.where(pd.notna(df), None)

        if '罚则' in df.columns:
            df = df.drop(columns='罚则')
        if '违法情形' in df.columns:
            df = df.drop(columns='违法情形')

        # 添加新的列
        df['罚则'] = 0
        df['违法情形'] = 0
        # 新增列：违法行为文本
        df['违法行为文本'] = None

        # 筛选出条值小于 10000 的记录
        df_less_than_10000 = df[df['条'] < 10000]
        df_greater_than_or_equal_to_10000 = df[df['条'] >= 10000]

        # 遍历条值小于 10000 的记录，设置罚则
        for index, row in df_less_than_10000.iterrows():
            content = str(row['内容'])  # 确保内容为字符串类型
            df_less_than_10000.at[index, '罚则'] = check_discretion(content)

        # 按条及条内的树进行遍历，设置违法情形
        for index, row in df_less_than_10000.iterrows():
            # 这里可以继续添加处理违法情形的代码，原代码中此处逻辑可继续沿用
            if row['罚则'] == 1:
                df_less_than_10000.at[index, '条类型'] = 2
                # 确保 row 中的缺失值已经被正确转换
                row = row.where(pd.notna(row), None)
                children = find_children(df_less_than_10000, row)
                if children.empty:
                    # 判断其是不是违法情形
                    associated_body = find_associated_body(row['内容'])
                    # 将 associated_body 写入“违法行为文本”列
                    df_less_than_10000.at[index, '违法行为文本'] = associated_body
                    pattern = re.compile(get_violation_pattern())
                    # 增加判断第一个逗号前是否为“的”的条件
                    comma_index = associated_body.find('，')
                    first_part_ends_with_de = comma_index > 0 and associated_body[comma_index - 1] == '的'
                    # 增加判断：associated_body最后一个的文字是不是“，”，同时判断其前面是不是“的”
                    last_comma_index = len(associated_body) - 1
                    last_part_ends_with_de = last_comma_index > 0 and associated_body[last_comma_index] == '，' and associated_body[last_comma_index - 1] == '的'
                    #print(f"associated_body1 {associated_body}")
                
                    if pattern.search(associated_body) or first_part_ends_with_de or last_part_ends_with_de:
                        df_less_than_10000.at[index, '违法情形'] = 1
                        df_less_than_10000.at[index, '条类型'] = 4
                elif  "下列" in row['内容']:
                    df_less_than_10000.at[index, '条类型'] = 6
                    for child_index, _ in children.iterrows():
                        df_less_than_10000.at[child_index, '违法情形'] = 1
                        df_less_than_10000.at[child_index, '条类型'] = 3

        # 合并处理后的记录和未处理的记录
        df = pd.concat([df_less_than_10000, df_greater_than_or_equal_to_10000]).sort_index()

        # 清空原工作表内容（除表头）
        for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.value = None

        # 写入表头
        headers = df.columns.tolist()
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # 写入数据
        for row_idx, row in df.iterrows():
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx + 2, column=col_idx, value=value)

    wb.save(file_path)



if __name__ == "__main__":
    file_path = "result\\law_structure_format_num_ex.xlsx"
    #file_path = "law_structure.xlsx"
    update_excel(file_path)
    print("文件更新完成。")